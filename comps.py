#!/usr/bin/env python3
"""Renewables Comparables Analysis — pulls live data via yfinance, exports to Excel."""

import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

SEGMENTS = {
    "Solar": ["ENPH", "FSLR", "RUN"],
    "Diversified Renewables": ["NEE", "BEP", "CWEN", "AES"],
    "Utilities": ["ED"],
    "Clean Energy Tech": ["PLUG", "BE"],
}

COL_KEYS = [
    "Ticker", "Market Cap", "Total Debt", "Cash",
    "Revenue", "EBITDA", "Net Income",
    "Shares Outstanding", "Enterprise Value",
    "EV/Revenue", "EV/EBITDA", "P/E",
]

COL_HEADERS = [
    "Ticker", "Market Cap ($M)", "Total Debt ($M)", "Cash ($M)",
    "Revenue ($M)", "EBITDA ($M)", "Net Income ($M)",
    "Shares Out. (M)", "Enterprise Value ($M)",
    "EV/Revenue (x)", "EV/EBITDA (x)", "P/E (x)",
]

MULTIPLE_KEYS = {"EV/Revenue", "EV/EBITDA", "P/E"}


def safe_get(info, key):
    val = info.get(key)
    return val if val is not None else None


def safe_div(numerator, denominator):
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return numerator / denominator


def to_millions(val):
    return val / 1_000_000 if val is not None else None


def fetch_ticker_data(ticker):
    t = yf.Ticker(ticker)
    info = t.info

    market_cap  = safe_get(info, "marketCap")
    total_debt  = safe_get(info, "totalDebt")
    cash        = safe_get(info, "totalCash")
    revenue     = safe_get(info, "totalRevenue")
    ebitda      = safe_get(info, "ebitda")
    net_income  = safe_get(info, "netIncomeToCommon")
    shares_out  = safe_get(info, "sharesOutstanding")

    # Enterprise Value = Market Cap + Total Debt − Cash
    if market_cap is not None:
        ev = market_cap + (total_debt or 0) - (cash or 0)
    else:
        ev = None

    ev_rev    = safe_div(ev, revenue)
    ev_ebitda = safe_div(ev, ebitda)
    pe        = safe_div(market_cap, net_income)

    # Console flags for missing / unusual values
    flags = []
    for label, val in [
        ("Market Cap", market_cap), ("Revenue", revenue),
        ("EBITDA", ebitda), ("Net Income", net_income),
    ]:
        if val is None:
            flags.append(f"{label} missing")
        elif val < 0:
            flags.append(f"{label} negative")
    if flags:
        print(f"    [!] {ticker}: {', '.join(flags)}")

    return {
        "Ticker":             ticker,
        "Market Cap":         market_cap,
        "Total Debt":         total_debt,
        "Cash":               cash,
        "Revenue":            revenue,
        "EBITDA":             ebitda,
        "Net Income":         net_income,
        "Shares Outstanding": shares_out,
        "Enterprise Value":   ev,
        "EV/Revenue":         ev_rev,
        "EV/EBITDA":          ev_ebitda,
        "P/E":                pe,
    }


def write_excel(all_data, output_path="renewables_comps.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comps"

    # --- Style constants ---
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    SEG_FILL     = PatternFill("solid", fgColor="BDD7EE")
    ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    NA_FONT      = Font(size=10, color="C00000", italic=True)
    TICKER_FONT  = Font(bold=True, size=10)
    DATA_FONT    = Font(size=10)
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SEG_FONT     = Font(bold=True, color="1F4E79", size=10)
    TITLE_FONT   = Font(bold=True, size=14, color="1F4E79")
    DATE_FONT    = Font(italic=True, size=9, color="808080")

    ncols = len(COL_KEYS)

    # --- Row 1: Title ---
    ws.append(["Renewables Comparables Analysis"] + [""] * (ncols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # --- Row 2: Date ---
    ws.append([f"As of {date.today().strftime('%B %d, %Y')}"] + [""] * (ncols - 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1).font = DATE_FONT
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 3: Blank spacer ---
    ws.append([""] * ncols)

    # --- Row 4: Column headers ---
    ws.append(COL_HEADERS)
    header_row = ws.max_row
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(header_row, col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 32

    # Track max content width per column for auto-sizing
    col_widths = [len(h) for h in COL_HEADERS]

    # --- Data rows ---
    data_row_count = 0
    for segment, tickers in SEGMENTS.items():
        # Segment subheader
        ws.append([segment] + [""] * (ncols - 1))
        seg_row = ws.max_row
        ws.merge_cells(start_row=seg_row, start_column=1, end_row=seg_row, end_column=ncols)
        seg_cell = ws.cell(seg_row, 1)
        seg_cell.font  = SEG_FONT
        seg_cell.fill  = SEG_FILL
        seg_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[seg_row].height = 18

        for ticker in tickers:
            row_data = all_data[ticker]
            data_row_count += 1
            fill = ALT_FILL if data_row_count % 2 == 0 else WHITE_FILL

            row_vals = []
            for key in COL_KEYS:
                raw = row_data[key]
                if key == "Ticker":
                    row_vals.append(raw)
                elif key in MULTIPLE_KEYS:
                    row_vals.append(raw)          # raw float or None
                else:
                    row_vals.append(to_millions(raw))  # convert to $M

            ws.append(row_vals)
            cur_row = ws.max_row
            ws.row_dimensions[cur_row].height = 16

            for col_idx, (key, val) in enumerate(zip(COL_KEYS, row_vals), 1):
                cell = ws.cell(cur_row, col_idx)
                cell.fill = fill

                if val is None:
                    cell.value     = "N/A"
                    cell.font      = NA_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], 5)
                elif key == "Ticker":
                    cell.font      = TICKER_FONT
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(val)))
                elif key in MULTIPLE_KEYS:
                    cell.font          = DATA_FONT
                    cell.number_format = '#,##0.00'
                    cell.alignment     = Alignment(horizontal="right", vertical="center")
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], 8)
                else:
                    cell.font          = DATA_FONT
                    cell.number_format = '#,##0.0'
                    cell.alignment     = Alignment(horizontal="right", vertical="center")
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], 12)

    # --- Auto-size columns ---
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 26)

    # Freeze header row and Ticker column
    ws.freeze_panes = "B5"

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


def main():
    print("Renewables Comparables Analysis")
    print("=" * 44)

    all_data = {}
    for segment, tickers in SEGMENTS.items():
        print(f"\n{segment}:")
        for ticker in tickers:
            print(f"  Fetching {ticker}...")
            try:
                all_data[ticker] = fetch_ticker_data(ticker)
            except Exception as exc:
                print(f"  ERROR {ticker}: {exc}")
                all_data[ticker] = {k: (ticker if k == "Ticker" else None) for k in COL_KEYS}

    write_excel(all_data)


if __name__ == "__main__":
    main()
